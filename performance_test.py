"""
ECSE-429 Part C – Performance Testing
Measures transaction times for create / delete / change operations
on the REST API Todo Manager as the number of objects grows.

:Author: Bilguun Tegshbayar

Usage:
    python performance_test.py

Requirements:
    pip install requests psutil matplotlib pandas openpyxl
"""

import time
import random
import string
import json
import csv
import os
import threading

import requests
import psutil
import matplotlib.pyplot as plt
import pandas as pd

BASE_URL = "http://localhost:4567"
TODOS_URL = f"{BASE_URL}/todos"
HEADERS = {"Content-Type": "application/json"}

# ── workload sizes to test ───────────────────────────────────────────────────
OBJECT_COUNTS = [10, 50, 100, 200, 500]
RESULTS_FILE = "performance_results.csv"

# ── resource snapshot helpers ────────────────────────────────────────────────

def snapshot_resources():
    """Return (cpu_percent, available_mb) at this instant."""
    cpu = psutil.cpu_percent(interval=0.1)
    mem = psutil.virtual_memory().available / (1024 * 1024)
    return cpu, mem


# ── resource monitor (background thread) ─────────────────────────────────────

class ResourceMonitor:
    """Collects CPU / memory samples in the background while an operation runs."""

    def __init__(self, interval=0.2):
        self.interval = interval
        self._cpus = []
        self._mems = []
        self._stop = threading.Event()

    def start(self):
        self._stop.clear()
        self._thread = threading.Thread(target=self._collect, daemon=True)
        self._thread.start()

    def stop(self):
        self._stop.set()
        self._thread.join()

    def _collect(self):
        while not self._stop.is_set():
            cpu, mem = snapshot_resources()
            self._cpus.append(cpu)
            self._mems.append(mem)
            time.sleep(self.interval)

    @property
    def avg_cpu(self):
        return sum(self._cpus) / len(self._cpus) if self._cpus else 0.0

    @property
    def avg_mem(self):
        return sum(self._mems) / len(self._mems) if self._mems else 0.0


# ── API helpers ───────────────────────────────────────────────────────────────

def random_string(length=8):
    return "".join(random.choices(string.ascii_letters, k=length))


def check_api_responsive():
    try:
        r = requests.get(TODOS_URL, timeout=5)
        return r.status_code == 200
    except requests.exceptions.ConnectionError:
        return False


def create_todo(title=None, description=None, done=False):
    payload = {
        "title": title or f"Task-{random_string()}",
        "description": description or random_string(20),
        "doneStatus": done,
    }
    r = requests.post(TODOS_URL, json=payload, headers=HEADERS)
    r.raise_for_status()
    return r.json()["id"]


def delete_todo(todo_id):
    r = requests.delete(f"{TODOS_URL}/{todo_id}")
    r.raise_for_status()


def update_todo(todo_id, new_title=None):
    payload = {"title": new_title or f"Updated-{random_string()}"}
    r = requests.put(f"{TODOS_URL}/{todo_id}", json=payload, headers=HEADERS)
    r.raise_for_status()


def get_all_todo_ids():
    r = requests.get(TODOS_URL)
    r.raise_for_status()
    return [t["id"] for t in r.json().get("todos", [])]


def delete_all_todos():
    for tid in get_all_todo_ids():
        try:
            requests.delete(f"{TODOS_URL}/{tid}")
        except Exception:
            pass


# ── experiment runner ─────────────────────────────────────────────────────────

def run_experiment(n: int):
    """
    Populate the DB with n todos, then measure:
      - time to add one more  (create)
      - time to delete one    (delete)
      - time to modify one    (update)
    Returns dict with timings and resource stats.
    """
    delete_all_todos()

    # populate with n items using random data
    ids = []
    for _ in range(n):
        ids.append(create_todo())

    # ── CREATE experiment ──────────────────────────────────────────────────
    monitor = ResourceMonitor()
    monitor.start()
    t0 = time.perf_counter()
    new_id = create_todo(title=f"Perf-Create-{n}")
    create_time = (time.perf_counter() - t0) * 1000  # ms
    monitor.stop()
    create_cpu = monitor.avg_cpu
    create_mem = monitor.avg_mem
    ids.append(new_id)

    # ── DELETE experiment ──────────────────────────────────────────────────
    target_id = ids[len(ids) // 2]  # pick from middle
    monitor = ResourceMonitor()
    monitor.start()
    t0 = time.perf_counter()
    delete_todo(target_id)
    delete_time = (time.perf_counter() - t0) * 1000
    monitor.stop()
    delete_cpu = monitor.avg_cpu
    delete_mem = monitor.avg_mem
    ids.remove(target_id)

    # ── UPDATE experiment ──────────────────────────────────────────────────
    target_id = ids[0]
    monitor = ResourceMonitor()
    monitor.start()
    t0 = time.perf_counter()
    update_todo(target_id, new_title=f"Updated-at-{n}")
    update_time = (time.perf_counter() - t0) * 1000
    monitor.stop()
    update_cpu = monitor.avg_cpu
    update_mem = monitor.avg_mem

    delete_all_todos()

    return {
        "n_objects": n,
        "create_ms": round(create_time, 3),
        "delete_ms": round(delete_time, 3),
        "update_ms": round(update_time, 3),
        "create_cpu_pct": round(create_cpu, 2),
        "delete_cpu_pct": round(delete_cpu, 2),
        "update_cpu_pct": round(update_cpu, 2),
        "create_mem_mb": round(create_mem, 1),
        "delete_mem_mb": round(delete_mem, 1),
        "update_mem_mb": round(update_mem, 1),
    }


# ── repeated runs for statistical stability ────────────────────────────────

def run_all_experiments(runs_per_size=3):
    all_rows = []
    for n in OBJECT_COUNTS:
        print(f"\n[Experiment] n={n} objects ({runs_per_size} runs) ...")
        for run in range(1, runs_per_size + 1):
            row = run_experiment(n)
            row["run"] = run
            all_rows.append(row)
            print(
                f"  run {run}: create={row['create_ms']}ms  "
                f"delete={row['delete_ms']}ms  update={row['update_ms']}ms"
            )
    return all_rows


# ── aggregate + save ───────────────────────────────────────────────────────

def aggregate(rows):
    df = pd.DataFrame(rows)
    agg = (
        df.groupby("n_objects")
        .agg(
            create_ms_mean=("create_ms", "mean"),
            create_ms_max=("create_ms", "max"),
            delete_ms_mean=("delete_ms", "mean"),
            delete_ms_max=("delete_ms", "max"),
            update_ms_mean=("update_ms", "mean"),
            update_ms_max=("update_ms", "max"),
            avg_cpu_pct=(
                "create_cpu_pct",
                lambda x: round(
                    (x.mean() + df.loc[x.index, "delete_cpu_pct"].mean() + df.loc[x.index, "update_cpu_pct"].mean()) / 3,
                    2,
                ),
            ),
            avg_mem_mb=(
                "create_mem_mb",
                lambda x: round(
                    (x.mean() + df.loc[x.index, "delete_mem_mb"].mean() + df.loc[x.index, "update_mem_mb"].mean()) / 3,
                    1,
                ),
            ),
        )
        .reset_index()
    )
    for col in ["create_ms_mean", "create_ms_max", "delete_ms_mean", "delete_ms_max", "update_ms_mean", "update_ms_max"]:
        agg[col] = agg[col].round(3)
    return df, agg


def save_csv(rows, path):
    if not rows:
        return
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    print(f"Raw results saved → {path}")


# ── chart generation ───────────────────────────────────────────────────────

def generate_charts(agg: pd.DataFrame, out_dir: str):
    os.makedirs(out_dir, exist_ok=True)
    x = agg["n_objects"]

    # 1. Transaction time chart
    fig, ax = plt.subplots(figsize=(10, 5))
    ax.plot(x, agg["create_ms_mean"], marker="o", label="Create (mean)")
    ax.plot(x, agg["delete_ms_mean"], marker="s", label="Delete (mean)")
    ax.plot(x, agg["update_ms_mean"], marker="^", label="Update (mean)")
    ax.fill_between(x, agg["create_ms_mean"], agg["create_ms_max"], alpha=0.15)
    ax.fill_between(x, agg["delete_ms_mean"], agg["delete_ms_max"], alpha=0.15)
    ax.fill_between(x, agg["update_ms_mean"], agg["update_ms_max"], alpha=0.15)
    ax.set_xlabel("Number of Objects in Database")
    ax.set_ylabel("Transaction Time (ms)")
    ax.set_title("Transaction Time vs. Number of Objects")
    ax.legend()
    ax.grid(True, linestyle="--", alpha=0.5)
    path1 = os.path.join(out_dir, "transaction_time.png")
    fig.savefig(path1, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"Chart saved → {path1}")

    # 2. CPU usage chart
    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(x, agg["avg_cpu_pct"], marker="D", color="tab:orange", label="Avg CPU %")
    ax.set_xlabel("Number of Objects in Database")
    ax.set_ylabel("CPU Usage (%)")
    ax.set_title("CPU Usage vs. Number of Objects")
    ax.legend()
    ax.grid(True, linestyle="--", alpha=0.5)
    path2 = os.path.join(out_dir, "cpu_usage.png")
    fig.savefig(path2, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"Chart saved → {path2}")

    # 3. Memory chart
    fig, ax = plt.subplots(figsize=(10, 4))
    ax.plot(x, agg["avg_mem_mb"], marker="v", color="tab:green", label="Available Memory (MB)")
    ax.set_xlabel("Number of Objects in Database")
    ax.set_ylabel("Available Memory (MB)")
    ax.set_title("Available Memory vs. Number of Objects")
    ax.legend()
    ax.grid(True, linestyle="--", alpha=0.5)
    path3 = os.path.join(out_dir, "memory_usage.png")
    fig.savefig(path3, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"Chart saved → {path3}")

    return path1, path2, path3


# ── Excel summary report ───────────────────────────────────────────────────

def generate_excel(raw_df: pd.DataFrame, agg: pd.DataFrame, out_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    alt_fill = PatternFill("solid", start_color="D6E4F0")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    cols = [
        "Objects", "Create (ms)", "Create Max (ms)",
        "Delete (ms)", "Delete Max (ms)",
        "Update (ms)", "Update Max (ms)",
        "Avg CPU (%)", "Avg Free Mem (MB)",
    ]
    ws.append(cols)
    for col_idx, _ in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for i, row in enumerate(agg.itertuples(), start=2):
        data = [
            row.n_objects,
            round(row.create_ms_mean, 3),
            round(row.create_ms_max, 3),
            round(row.delete_ms_mean, 3),
            round(row.delete_ms_max, 3),
            round(row.update_ms_mean, 3),
            round(row.update_ms_max, 3),
            round(row.avg_cpu_pct, 2),
            round(row.avg_mem_mb, 1),
        ]
        ws.append(data)
        fill = alt_fill if i % 2 == 0 else PatternFill()
        for col_idx in range(1, len(cols) + 1):
            cell = ws.cell(row=i, column=col_idx)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 18

    ws.freeze_panes = "A2"

    # ── Sheet 2: Raw Data ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Raw Data")
    raw_cols = list(raw_df.columns)
    ws2.append(raw_cols)
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in raw_df.itertuples(index=False):
        ws2.append(list(row))

    for col in ws2.columns:
        ws2.column_dimensions[get_column_letter(col[0].column)].width = 16

    # ── Sheet 3: Charts ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Charts")
    chart_dir = os.path.join(os.path.dirname(out_path), "charts")
    chart_files = ["transaction_time.png", "cpu_usage.png", "memory_usage.png"]
    titles = ["Transaction Time", "CPU Usage", "Available Memory"]
    row_offset = 2
    for fname, title in zip(chart_files, titles):
        fpath = os.path.join(chart_dir, fname)
        if os.path.exists(fpath):
            ws3.cell(row=row_offset, column=1, value=title).font = Font(bold=True, size=12)
            img = XLImage(fpath)
            img.width = 700
            img.height = 320
            ws3.add_image(img, f"A{row_offset + 1}")
            row_offset += 22

    wb.save(out_path)
    print(f"Excel report saved → {out_path}")


# ── entry point ────────────────────────────────────────────────────────────

def main():
    print("ECSE-429 Part C – Performance Testing")
    print("=" * 50)

    if not check_api_responsive():
        raise SystemExit(
            "\n[ERROR] API is not running at http://localhost:4567\n"
            "Start it with:  java -jar runTodoManagerRestAPI-1.5.2.jar"
        )

    print(f"API confirmed responsive. Running experiments for n ∈ {OBJECT_COUNTS} ...\n")

    out_dir = os.path.dirname(os.path.abspath(__file__))
    chart_dir = os.path.join(out_dir, "charts")

    rows = run_all_experiments(runs_per_size=3)
    raw_df, agg = aggregate(rows)

    save_csv(rows, os.path.join(out_dir, RESULTS_FILE))
    generate_charts(agg, chart_dir)
    generate_excel(raw_df, agg, os.path.join(out_dir, "performance_report.xlsx"))

    print("\n── Aggregated Summary ──────────────────────────────────")
    print(agg.to_string(index=False))
    print("\nDone.")


if __name__ == "__main__":
    main()
