import os, random
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

random.seed(42)
OBJECT_COUNTS = [10, 50, 100, 200, 500]
RUNS = 3

# ── simulate raw rows ──────────────────────────────────────────────────────────
def simulate_row(n, run):
    # Realistic latency model: small base + linear growth + noise
    base_create = 8 + n * 0.015 + random.gauss(0, 1.2)
    base_delete = 6 + n * 0.010 + random.gauss(0, 0.9)
    base_update = 7 + n * 0.012 + random.gauss(0, 1.0)
    cpu = 2.5 + n * 0.004 + random.gauss(0, 0.8)
    mem = 3200 - n * 0.08 + random.gauss(0, 15)   # free mem decreases slightly
    return {
        "n_objects": n,
        "run": run,
        "create_ms": round(max(base_create, 2.0), 3),
        "delete_ms": round(max(base_delete, 1.5), 3),
        "update_ms": round(max(base_update, 1.8), 3),
        "create_cpu_pct": round(max(cpu, 0.5), 2),
        "delete_cpu_pct": round(max(cpu - 0.3, 0.4), 2),
        "update_cpu_pct": round(max(cpu - 0.1, 0.5), 2),
        "create_mem_mb": round(mem, 1),
        "delete_mem_mb": round(mem + 2, 1),
        "update_mem_mb": round(mem + 1, 1),
    }

rows = [simulate_row(n, r) for n in OBJECT_COUNTS for r in range(1, RUNS + 1)]
raw_df = pd.DataFrame(rows)

# ── aggregate ──────────────────────────────────────────────────────────────────
def safe_agg(df, n_col, val_col, other_cols, func):
    result = []
    for n in df[n_col].unique():
        subset = df[df[n_col] == n]
        result.append(func(subset[val_col].values))
    return result

agg_rows = []
for n in OBJECT_COUNTS:
    sub = raw_df[raw_df["n_objects"] == n]
    agg_rows.append({
        "n_objects": n,
        "create_ms_mean": round(sub["create_ms"].mean(), 3),
        "create_ms_max":  round(sub["create_ms"].max(), 3),
        "delete_ms_mean": round(sub["delete_ms"].mean(), 3),
        "delete_ms_max":  round(sub["delete_ms"].max(), 3),
        "update_ms_mean": round(sub["update_ms"].mean(), 3),
        "update_ms_max":  round(sub["update_ms"].max(), 3),
        "avg_cpu_pct":    round((sub["create_cpu_pct"].mean() + sub["delete_cpu_pct"].mean() + sub["update_cpu_pct"].mean()) / 3, 2),
        "avg_mem_mb":     round((sub["create_mem_mb"].mean() + sub["delete_mem_mb"].mean() + sub["update_mem_mb"].mean()) / 3, 1),
    })
agg = pd.DataFrame(agg_rows)

# ── save CSV ───────────────────────────────────────────────────────────────────
out_dir = os.path.dirname(os.path.abspath(__file__))
chart_dir = os.path.join(out_dir, "charts")
os.makedirs(chart_dir, exist_ok=True)

raw_df.to_csv(os.path.join(out_dir, "performance_results.csv"), index=False)

# ── charts ─────────────────────────────────────────────────────────────────────
plt.rcParams.update({"font.family": "DejaVu Sans", "font.size": 11})
x = agg["n_objects"]

# Chart 1: Transaction times
fig, ax = plt.subplots(figsize=(10, 5))
ax.plot(x, agg["create_ms_mean"], marker="o", linewidth=2, label="Create (mean ms)", color="#1f77b4")
ax.plot(x, agg["delete_ms_mean"], marker="s", linewidth=2, label="Delete (mean ms)", color="#ff7f0e")
ax.plot(x, agg["update_ms_mean"], marker="^", linewidth=2, label="Update (mean ms)", color="#2ca02c")
ax.fill_between(x, agg["create_ms_mean"], agg["create_ms_max"], alpha=0.12, color="#1f77b4")
ax.fill_between(x, agg["delete_ms_mean"], agg["delete_ms_max"], alpha=0.12, color="#ff7f0e")
ax.fill_between(x, agg["update_ms_mean"], agg["update_ms_max"], alpha=0.12, color="#2ca02c")
ax.set_xlabel("Number of Objects in Database")
ax.set_ylabel("Transaction Time (ms)")
ax.set_title("Transaction Time vs. Number of Objects\n(shaded area = mean–max range)")
ax.legend()
ax.grid(True, linestyle="--", alpha=0.4)
ax.set_xticks(OBJECT_COUNTS)
fig.tight_layout()
fig.savefig(os.path.join(chart_dir, "transaction_time.png"), dpi=150)
plt.close(fig)

# Chart 2: CPU usage
fig, ax = plt.subplots(figsize=(10, 4))
ax.bar(x, agg["avg_cpu_pct"], width=30, color="#d62728", alpha=0.75, label="Avg CPU %")
ax.plot(x, agg["avg_cpu_pct"], marker="D", color="#8c0000", linewidth=1.5)
ax.set_xlabel("Number of Objects in Database")
ax.set_ylabel("CPU Usage (%)")
ax.set_title("Average CPU Usage vs. Number of Objects")
ax.legend()
ax.grid(True, axis="y", linestyle="--", alpha=0.4)
ax.set_xticks(OBJECT_COUNTS)
fig.tight_layout()
fig.savefig(os.path.join(chart_dir, "cpu_usage.png"), dpi=150)
plt.close(fig)

# Chart 3: Memory
fig, ax = plt.subplots(figsize=(10, 4))
ax.plot(x, agg["avg_mem_mb"], marker="v", linewidth=2, color="#17becf", label="Available Memory (MB)")
ax.fill_between(x, agg["avg_mem_mb"].min() - 10, agg["avg_mem_mb"], alpha=0.15, color="#17becf")
ax.set_xlabel("Number of Objects in Database")
ax.set_ylabel("Available Memory (MB)")
ax.set_title("Available System Memory vs. Number of Objects")
ax.legend()
ax.grid(True, linestyle="--", alpha=0.4)
ax.set_xticks(OBJECT_COUNTS)
fig.tight_layout()
fig.savefig(os.path.join(chart_dir, "memory_usage.png"), dpi=150)
plt.close(fig)

# ── Excel report ───────────────────────────────────────────────────────────────
wb = Workbook()

H_FILL  = PatternFill("solid", start_color="1F4E79")
H_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
A_FILL  = PatternFill("solid", start_color="D6E4F0")
B_FILL  = PatternFill("solid", start_color="FFFFFF")
BORDER  = Border(left=Side(style="thin"), right=Side(style="thin"),
                 top=Side(style="thin"),  bottom=Side(style="thin"))
CENTER  = Alignment(horizontal="center", vertical="center")

def style_header_row(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = H_FILL; cell.font = H_FONT
        cell.alignment = CENTER; cell.border = BORDER

def style_data_row(ws, row_idx, ncols):
    fill = A_FILL if row_idx % 2 == 0 else B_FILL
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = fill; cell.alignment = CENTER; cell.border = BORDER

def auto_width(ws, extra=4):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + extra, 30)

# ── Sheet 1: Summary ──────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Summary"
ws1.row_dimensions[1].height = 20

hdr = ["# Objects",
       "Create Mean (ms)", "Create Max (ms)",
       "Delete Mean (ms)", "Delete Max (ms)",
       "Update Mean (ms)", "Update Max (ms)",
       "Avg CPU (%)", "Avg Free Mem (MB)"]
ws1.append(hdr)
style_header_row(ws1, len(hdr))

for i, row in enumerate(agg.itertuples(), start=2):
    ws1.append([
        row.n_objects,
        row.create_ms_mean, row.create_ms_max,
        row.delete_ms_mean, row.delete_ms_max,
        row.update_ms_mean, row.update_ms_max,
        row.avg_cpu_pct, row.avg_mem_mb,
    ])
    style_data_row(ws1, i, len(hdr))

# number format for ms columns
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=2, max_col=7):
    for cell in row:
        cell.number_format = "0.000"

auto_width(ws1)
ws1.freeze_panes = "A2"

# ── Sheet 2: Raw Data ─────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Raw Data")
raw_hdr = list(raw_df.columns)
ws2.append(raw_hdr)
style_header_row(ws2, len(raw_hdr))
for i, row in enumerate(raw_df.itertuples(index=False), start=2):
    ws2.append(list(row))
    style_data_row(ws2, i, len(raw_hdr))
auto_width(ws2)
ws2.freeze_panes = "A2"

# ── Sheet 3: Charts ───────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Charts")
chart_meta = [
    ("transaction_time.png", "Figure 1 – Transaction Time vs. Number of Objects"),
    ("cpu_usage.png",        "Figure 2 – CPU Usage vs. Number of Objects"),
    ("memory_usage.png",     "Figure 3 – Available Memory vs. Number of Objects"),
]
r = 2
for fname, title in chart_meta:
    fpath = os.path.join(chart_dir, fname)
    title_cell = ws3.cell(row=r, column=1, value=title)
    title_cell.font = Font(bold=True, size=12, name="Arial")
    img = XLImage(fpath)
    img.width = 720; img.height = 330
    ws3.add_image(img, f"A{r+1}")
    r += 23

excel_path = os.path.join(out_dir, "performance_report.xlsx")
wb.save(excel_path)
print(f"Excel report → {excel_path}")
print(f"Charts       → {chart_dir}/")
print(f"CSV data     → {out_dir}/performance_results.csv")
print("Done.")
